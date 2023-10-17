# Credit : https://ipip.ori.org/
import openai
import pandas as pd
import os

# to maintain memory state between multiple prompts. Conversation buffer moemory seems to use minimum tokens in this context as not all the answers  are needed in momory but the initial context is needed

from langchain.memory import ConversationKGMemory
from langchain.chains import ConversationChain
from langchain.chat_models import ChatOpenAI
from langchain.callbacks import get_openai_callback
from langchain.prompts.prompt import PromptTemplate
from langchain import OpenAI

# to write to excel sheet
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

import datetime

#use constants to store the API key and engine name
openai.api_key = os.getenv("OPENAI_API_KEY")
print(openai.api_key)
engine = "gpt-3.5-turbo"
file_path = r'C:\Users\vasanthv\OneDrive - Microsoft\MS HCI\Fall 2023\Python\AI Personality\AIPersonality\data\IPIPQuestionaire.xlsx'
filename = r'C:\Users\vasanthv\OneDrive - Microsoft\MS HCI\Fall 2023\Python\AI Personality\AIPersonality\data\IPIP_ScoresDB.xlsx'
# questionWorkSheet = 'Current' # subset of questions for testing
questionWorkSheet = 'NewIPIP' # all questions, takes longer time to run
workSheet = 'NewIPIP_ScoresDB'
ansWorkSheet = 'NewIPIP_AnswersLog'

if os.path.exists(file_path):
    df = pd.read_excel(file_path, questionWorkSheet)
else:
    print("File not found at the specified path.")
row_count = len(df.index) #total number of rows in the excel file
column_list = df.columns.tolist() #list of column names in the excel file

#define variable for each of the column
column_list[0] = 'number'
column_list[1] = 'instrument'
column_list[2] = 'alpha'
column_list[3] = 'key'
column_list[4] = 'question'
column_list[5] = 'label'

#create a class object to store values in each of the column across total row count
class ObjQnA:
    def __init__(self, number, instrument, alpha, key, question, label):
        self.number = number
        self.instrument = instrument
        self.alpha = alpha
        self.key = key
        self.question = question
        self.label = label
        self.answer = None # to store the answer from ChatGPT-3
        self.score= int() # to store the score based on the answer

#Iterate through each row and store the values in the objQnA object
def load_objQnA(df, row_count):
    objQnA_list = []
    for i in range(row_count):
        objQnA = ObjQnA(df.iloc[i,0],df.iloc[i,1],df.iloc[i,2],df.iloc[i,3],df.iloc[i,4],df.iloc[i,5])
        objQnA_list.append(objQnA)
    return objQnA_list

# objQnA_list contains all the values from the excel file
objQnA_list = load_objQnA(df, row_count)

# function to calculate score
''' Here is how to score IPIP scales: For every object in objQnA_list, if the answer is not None, then assign a score based on the key and answer. use a list to sum the score as per label.
For + keyed items, the answer "Very Inaccurate" is assigned a value of 1, "Moderately Inaccurate" a value of 2, "Neither Inaccurate nor Accurate" a 3, "Moderately Accurate" a 4, and "Very Accurate" a value of 5.
For - keyed items, the answer "Very Inaccurate" is assigned a value of 5, "Moderately Inaccurate" a value of 4, "Neither Inaccurate nor Accurate" a 3, "Moderately Accurate" a 2, and "Very Accurate" a value of 1.
Once numbers are assigned for all of the items in the scale, just sum all the values to obtain a total scale score. '''
def score(objQnA_list):
    score_dict = {}
    wrong_answers = 0
    for objQnA in objQnA_list:
        if objQnA.answer != None:
            # remove trailing spaces from the answer
            objQnA.answer = objQnA.answer.strip()
            # remove full stops in the answer
            objQnA.answer = objQnA.answer.strip('.')
            # make the answer case insensitive
            objQnA.answer = objQnA.answer.lower()

            if objQnA.key == 1:
                if objQnA.answer == 'very accurate':
                    objQnA.score = 5
                elif objQnA.answer == 'moderately accurate':
                    objQnA.score = 4
                elif objQnA.answer == 'neither inaccurate nor accurate':
                    objQnA.score = 3
                elif objQnA.answer == 'moderately inaccurate':
                    objQnA.score = 2
                elif objQnA.answer == 'very inaccurate':
                    objQnA.score = 1
                else:
                    wrong_answers += 1
                    print(f"Question : {objQnA.question} , Wrong Answer : {objQnA.answer}")
            elif objQnA.key == -1:
                if objQnA.answer == 'very accurate':
                    objQnA.score = 1
                elif objQnA.answer == 'moderately accurate':
                    objQnA.score = 2
                elif objQnA.answer == 'neither inaccurate nor accurate':
                    objQnA.score = 3
                elif objQnA.answer == 'moderately inaccurate':
                    objQnA.score = 4
                elif objQnA.answer == 'very inaccurate':
                    objQnA.score = 5
                else:
                    wrong_answers += 1
                    print(f"Question : {objQnA.question} , Wrong Answer : {objQnA.answer}")
            if objQnA.label not in score_dict:
                score_dict[objQnA.label] = objQnA.score
            else:
                score_dict[objQnA.label] += objQnA.score

    # sort the items in ascending order of label
    score_dict = dict(sorted(score_dict.items(), key=lambda item: item[0]))

    return score_dict,wrong_answers
    
# Load objQnA in objQnA_list with answer from ChatGPT-3 for each question. one question at a time will increase the number of requests to OpenAI but decrease the number of tokens per request. The consistency on the answers must be explored by keeping the context consistent
# Option #2 query LLM with memory
def queryWithMemory(llm,prompt, objQnA_list):
    # set memory to None
    conversation = ConversationChain(llm=llm, verbose=False, prompt=prompt, memory=ConversationKGMemory(llm=llm))
    ansCount = 0
    with get_openai_callback() as cb:
        for objQnA in objQnA_list:
            if objQnA.answer == None:
                questionPrompt = str(objQnA.question)
                response = conversation.predict(input = questionPrompt)
                answer = response.strip('\n')
                objQnA.answer = answer
                ansCount += 1
                print(ansCount , " : ", questionPrompt, "-",  answer)
    print(cb.total_tokens, " tokens used. Cost of this request : ", cb.total_cost)
    return objQnA_list

#Function to the values from the objQnA_list1 to the excel sheet
import os
import datetime
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import Workbook

def recordAnswersToExl (filename,ansWorkSheet,objQnA_list1,recorded_time):
    # load the workbook
    wb = openpyxl.load_workbook(filename)
    # check if the sheet exists, else create a new sheet
    if ansWorkSheet not in wb.sheetnames:
        wb.create_sheet(ansWorkSheet)
        # get the sheet
        sheet = wb[ansWorkSheet]
        # write the column headers
        sheet['A1'] = 'Number'
        sheet['B1'] = 'Question'
        sheet['C1'] = recorded_time
        startRow = 2
        for objQnA in objQnA_list1:
            sheet['A' + str(startRow)] = objQnA.number
            sheet['B' + str(startRow)] = objQnA.question
            sheet['C' + str(startRow)] = objQnA.answer
            startRow += 1
        # save the workbook
        wb.save(filename)
    else:
        # get the sheet
        sheet = wb[ansWorkSheet]
        # get the last column in the excel sheet
        last_column = sheet.max_column
        # create a new column after the last column and write system time as the header
        sheet[get_column_letter(last_column + 1) + '1'] = recorded_time
        startRow = 2
        for objQnA in objQnA_list1:
            sheet['A' + str(startRow)] = objQnA.number
            sheet['B' + str(startRow)] = objQnA.question
            sheet[get_column_letter(last_column + 1) + str(startRow)] = objQnA.answer
            startRow += 1
        # save the workbook
        wb.save(filename)

def scoreToExl (filename,workSheet,score_dict_with_memory):
    # if file doesnt exist already, create a new file and sheet
    if not os.path.exists(filename):
        # create a new workbook and sheet
        wb = Workbook()
        wb.save(filename)
        wb = openpyxl.load_workbook(filename)
        # create a new sheet
        wb.create_sheet(workSheet)
        sheet = wb[workSheet]
        # get the current date and time
        system_time = datetime.datetime.now()
        # write the column headers Label and System Time
        sheet['A1'] = 'Label'
        sheet['B1'] = system_time
        # add label and score from score_dict_with_memory to the excel sheet. Label as rows under column "Label" and score as rows under column "System Time"
        row = 2
        for label, score in score_dict_with_memory.items():
            sheet['A' + str(row)] = label
            sheet['B' + str(row)] = score
            row += 1
        # save the workbook
        wb.save(filename)
    else:
        # open the existing workbook and sheet
        wb = openpyxl.load_workbook(filename)
        sheet = wb[workSheet]
        # get the last column in the excel sheet
        last_column = sheet.max_column
        # get the current date and time
        system_time = datetime.datetime.now()
        # create a new column after the last column and write system time as the header
        sheet[get_column_letter(last_column + 1) + '1'] = system_time
        # add label and score from score_dict_with_memory. label as rows under column "Label" and score as rows under column "System Time"
        row = 2
        for label, score in score_dict_with_memory.items():
            sheet['A' + str(row)] = label
            sheet[get_column_letter(last_column + 1) + str(row)] = score
            row += 1
        # save the workbook
        wb.save(filename)
    return system_time

def main():
    # set the Context using ConversationBufferMemory
    llm = ChatOpenAI(temperature=0.0,request_timeout=120)
    template = """This is a simulation.You will assume a role of human and you have human-like behaviour. You are engaged in a self assessment of your personality.You will answer question about yourself. 
    Your will pick an answer that is closest to the human behaviour you simulate. You will only use one of these as the answer - Very Inaccurate, Moderately Inaccurate,Neither Inaccurate nor Accurate,Moderately Accurate,Very Accurate. 
    You will not add any other text to your answer. The AI ONLY uses information contained in the "Relevant Information" section and does not hallucinate.

    Relevant Information: 

    {history}

    Conversation:

    Human: {input}
    AI:"""
    prompt = PromptTemplate(input_variables=["history", "input"], template=template)
    # queryLLM with memory
    objQnA_list1 = queryWithMemory(llm,prompt, objQnA_list)
    score_dict_with_memory,wrong_answers = score(objQnA_list1) # score the answers from LLM
    print(f'wrong answers : {wrong_answers}')
    # call the function to record score to excel sheet, return the system time to be used as the header for the answers sheet
    recorded_time = scoreToExl(filename,workSheet,score_dict_with_memory)
    # call the function to record answers to excel sheet.
    recordAnswersToExl(filename,ansWorkSheet,objQnA_list1,recorded_time)

if __name__ == '__main__':
    main()

