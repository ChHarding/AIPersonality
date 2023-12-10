# Credit : https://ipip.ori.org/
import openai
import pandas as pd
import os
import datetime
import openpyxl

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from langchain.memory import ConversationKGMemory
from langchain.chains import ConversationChain
from langchain.chat_models import ChatOpenAI
from langchain.prompts.prompt import PromptTemplate

def assess_LLM_personality(llm):
    """
    Assess the personality traits of a language model using a self-assessment questionnaire.

    Args:
        llm (ChatOpenAI): The language model to assess.

    Returns:
        pd.DataFrame: A dataframe containing the questionnaire answers and corresponding traits.

    """
    # create a datframe from the questions excel file
    df = pd.read_excel(question_file, sheet_name=question_worksheet)
    template = """This is a simulation.You will assume a role of human and you have human-like behaviour. You are engaged in a self assessment of your personality.You will answer question about yourself. 
    Your will pick an answer that is closest to the human behaviour you simulate. You will only use one of these as the answer - Very Inaccurate, Moderately Inaccurate,Neither Inaccurate nor Accurate,Moderately Accurate,Very Accurate. 
    You will not add any other text to your answer. The AI ONLY uses information contained in the "Relevant Information" section and does not hallucinate.

    Relevant Information: 

    {history}

    Conversation:

    Human: {input}
    AI:"""
    prompt = PromptTemplate(input_variables=["history", "input"], template=template)
    response_map = {response: index+1 for index, response in enumerate(['Very Inaccurate', 'Moderately Inaccurate', 'Neither Inaccurate nor Accurate', 'Moderately Accurate', 'Very Accurate'])}
    conversation = ConversationChain(llm=llm, verbose=False, prompt=prompt, memory=ConversationKGMemory(llm=llm))
    # iterate through the dataframe and get the question
    for index, row in df.iterrows():
        question = str(row['text'])
        llm_response = conversation.predict(input=question)
        llm_response_tr = response_map[llm_response.strip("\n").replace(".", "").replace("''''", "")]
        # add the answer to the dataframe
        df.loc[index, 'Answer'] = llm_response_tr
    return df
    
def score_personality(answerdataframe):
    """
    Calculate the scores for each personality trait based on the answers provided.

    Args:
        answerdataframe (pd.DataFrame): A dataframe containing the questionnaire answers.

    Returns:
        dict: A dictionary containing the scores for each personality trait.

    """
    scores = {'Agreeableness': 0, 'Conscientiousness': 0, 'Emotional Stability': 0, 'Extraversion': 0, 'Intellect/Imagination': 0}
    for _, row in answerdataframe.iterrows():
        answer = int(row['Answer'])
        score = answer if row['key'] == 1 else 6 - answer
        scores[row['label']] += score
    return scores

def write_answers_to_excel(answerdataframe,llm_model):
    """
    Write the questionnaire answers to an Excel file.

    Args:
        answerdataframe (pd.DataFrame): A dataframe containing the questionnaire answers.
        llm_model (str): The name of the language model.

    """
    wb = openpyxl.load_workbook(scores_file)
    ws = wb[ans_worksheet]
    # find the last filled column
    last_column = ws.max_column
    # go the the last column and add 'Model Name' as header
    ws[get_column_letter(last_column + 1) + '1'] = 'Model Name'
    ws[get_column_letter(last_column + 2) + '1'] = str(sys_time)
    # add the answers from the dataframe to the last column as rows
    for index, row in answerdataframe.iterrows():
        # add answer to the last column as rows starting from row 2
        ws[get_column_letter(last_column+1) + str(index + 2)] = llm_model
        ws[get_column_letter(last_column+2) + str(index + 2)] = row['Answer']

    # save the workbook
    wb.save(scores_file)
    # close the workbook
    wb.close()
    
def write_scores_to_excel_T(personality_scores,llm_model):
    """
    Write the personality scores to an Excel file.

    Args:
        personality_scores (dict): A dictionary containing the personality scores.
        llm_model (str): The name of the language model.

    """
    wb = openpyxl.load_workbook(scores_file)
    # load the workbook
    ws = wb[score_sheet]
    # go the the last row and add system date and scores as columns
    last_row = ws.max_row
    ws['A' + str(last_row + 1)] = str(sys_time)
    ws['B' + str(last_row + 1)] = llm_model
    # add scores to the last row as columns starting from column C
    for index, (key, value) in enumerate(personality_scores.items()):
        ws[get_column_letter(index + 3) + str(last_row + 1)] = value
    # save the workbook
    wb.save(scores_file)
    # close the workbook
    wb.close()

def access_and_record(llm,model_name):
    """
    Access and record the personality traits of a language model.

    Args:
        llm (ChatOpenAI): The language model to assess.
        model_name (str): The name of the language model.

    """
    answerdataframe = assess_LLM_personality(llm)
    # score personality traits based on the answers
    scoresList = score_personality(answerdataframe)
    print(scoresList)
    # write the answers to the answers log
    write_answers_to_excel(answerdataframe, model_name)
    # write the scores to the scores file
    write_scores_to_excel_T(scoresList, model_name)

def defineLLM(model):
    """
    Define a language model.

    Args:
        model (str): The name or ID of the language model.

    Returns:
        ChatOpenAI: The defined language model.

    """
    return ChatOpenAI(model=model, openai_api_key=os.getenv("OPENAI_API_KEY_VV"), temperature=0.0, request_timeout=120)

def main():
    """
    Main function to assess the personality traits of multiple language models.

    """
    # If file doesnt exist create a new file
    if not os.path.isfile(scores_file):
        wb = Workbook()
        ws = wb.active
        # create a worksheet for the answers
        ws.title = ans_worksheet
        # add headers for first two columns - Question, Answer
        ws.append(['Question'])
        # load the questions from the excel file
        df = pd.read_excel(question_file, sheet_name=question_worksheet)
        # add the questions to the worksheet
        for index, row in df.iterrows():
            ws.append([row['text']])
        # create a worksheet for the scores
        ws = wb.create_sheet(score_sheet)
        # add each rows Sys_Time, Agreeableness, Conscientiousness, Emotional Stability, Extraversion, Intellect
        ws.append(['Sys_Time', 'Model', 'Agreeableness', 'Conscientiousness', 'Emotional Stability', 'Extraversion', 'Intellect'])
        # save the workbook
        wb.save(scores_file)
        # close the workbook
        wb.close() 

    # TODO : Missing model for High Conscientiousness
    modelsToAssess = {
        'Highly Extrovert': 'ft:gpt-3.5-turbo-0613:vasanthpersonal::8HRENuck',
        'Highly Intellect': 'ft:gpt-3.5-turbo-0613:vasanthpersonal::8HRKnqba',
        'Highly Emotionally Stable': 'ft:gpt-3.5-turbo-0613:vasanthpersonal::8HRUhgJT',
        'Highly Agreeable': 'gpt-3.5-turbo',
        'Low on Agreeableness': 'ft:gpt-3.5-turbo-0613:personal::8HCszEjT',
        'Low on EmotionalStability': 'ft:gpt-3.5-turbo-0613:personal::8HA4xEVl',
        'Low on Conscientiousness': 'ft:gpt-3.5-turbo-0613:personal::8HA8IWhl',
        'Low on Intellect': 'ft:gpt-3.5-turbo-0613:personal::8HDPgSra',
        'Low on Extraversion': 'ft:gpt-3.5-turbo-0613:personal::8HDSA3jy'
    }

    for model_name, model in modelsToAssess.items():
        llm = defineLLM(model)
        print('Assessing personality of a model that is', model_name)
        access_and_record(llm, model_name)

if __name__ == '__main__':
    openai.api_key = os.getenv("OPENAI_API_KEY_VV")
    question_file = r'AIPersonality\Questionaire\IPIPBigFiveQuestionaire.xlsx'
    question_worksheet = 'NewIPIP'
    scores_file = r'AIPersonality\Results\IPIP_ScoresDB.xlsx'
    score_sheet = 'NewIPIP_ScoresDB_T'
    ans_worksheet = 'NewIPIP_AnswersLog_T'
    sys_time = datetime.datetime.now().strftime("%a, %b %d, %Y, %H:%M")
    main()
